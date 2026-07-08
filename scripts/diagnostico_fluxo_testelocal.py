from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
TESTELOCAL = BASE_DIR / "testelocal"
GABARITO = TESTELOCAL / "GABARITO.xlsx"
RESULTADO = TESTELOCAL / "RESULTADO.xlsx"
REPORT = TESTELOCAL / "diagnostico_fluxo.md"


CRITICAL_CELLS = {
    "DENS": [
        "B4",
        "F4",
        "K4",
        "M4",
        "B5",
        "K6",
        "M6",
        "B8",
        "F8",
        "J8",
        "G10",
        "G11",
        "G12",
        "G13",
        "G14",
        "C23",
        "D23",
        "L23",
        "N23",
        "C24",
        "D24",
        "L24",
        "N24",
        "M46",
        "M53",
    ],
    "CBR": [
        "B4",
        "F4",
        "L4",
        "O4",
        "I6",
        "M6",
        "O6",
        "B8",
        "H8",
        "L8",
        "H11",
        "H12",
        "H13",
        "H14",
        "H15",
        "H16",
        "H17",
        "H18",
        "O11",
        "O12",
        "O13",
        "O14",
        "O15",
        "O17",
        "P11",
        "P12",
        "P13",
        "P14",
        "P15",
        "P17",
        "F21",
        "F22",
        "F23",
        "M20",
        "M21",
        "O21",
        "P36",
        "P38",
        "P40",
        "P42",
        "F29",
        "G29",
        "J29",
        "F30",
        "G30",
        "J30",
        "F34",
        "N33",
        "P33",
        "J34",
        "P45",
        "P47",
        "P49",
        "P51",
    ],
}


def cell_value(workbook, sheet: str, cell: str) -> Any:
    if sheet not in workbook.sheetnames:
        return "<aba ausente>"
    return workbook[sheet][cell].value


def fmt(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def compare_workbooks() -> list[str]:
    lines: list[str] = [
        "# Diagnostico do fluxo testelocal",
        "",
        "Este relatorio compara celulas criticas entre `GABARITO.xlsx` e `RESULTADO.xlsx`.",
        "",
    ]
    if not GABARITO.exists() or not RESULTADO.exists():
        lines.append("Arquivos GABARITO.xlsx ou RESULTADO.xlsx nao encontrados.")
        return lines

    gabarito = load_workbook(GABARITO, data_only=False)
    resultado = load_workbook(RESULTADO, data_only=False)

    for sheet, cells in CRITICAL_CELLS.items():
        lines.extend([
            f"## Aba {sheet}",
            "",
            "| Celula | Gabarito | Resultado | Status |",
            "|---|---|---|---|",
        ])
        for cell in cells:
            expected = cell_value(gabarito, sheet, cell)
            actual = cell_value(resultado, sheet, cell)
            status = "OK" if expected == actual else "DIFERENTE"
            lines.append(f"| {cell} | {fmt(expected)} | {fmt(actual)} | {status} |")
        lines.append("")
    lines.extend([
        "## Leitura",
        "",
        "- `DIFERENTE` nao significa necessariamente erro: pode indicar que o gabarito e o resultado usam fontes diferentes.",
        "- O objetivo e localizar rapidamente onde o fluxo antigo divergiu para orientar a revisao humana e os testes.",
        "- Este diagnostico nao chama a OpenAI API.",
        "",
    ])
    return lines


def main() -> None:
    REPORT.write_text("\n".join(compare_workbooks()), encoding="utf-8")
    print(REPORT)


if __name__ == "__main__":
    main()

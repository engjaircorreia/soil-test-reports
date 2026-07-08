from django.test import override_settings
from openpyxl import load_workbook

from ensaios.services.validation import validate_workbook
from ensaios.services.workbook_fill import generate_workbooks


def sample_compaction_points():
    return [
        {
            "ponto": 1,
            "peso_umido_molde_g": 9374,
            "peso_solo_umido_g": 3960,
            "densidade_umida_g_cm3": 1.915,
            "umidade_media_percent": 3.5,
            "densidade_seca_g_cm3": 1.85,
        },
        {
            "ponto": 2,
            "peso_umido_molde_g": 9613,
            "peso_solo_umido_g": 4199,
            "densidade_umida_g_cm3": 2.03,
            "umidade_media_percent": 5.0,
            "densidade_seca_g_cm3": 1.933,
        },
        {
            "ponto": 3,
            "peso_umido_molde_g": 9874,
            "peso_solo_umido_g": 4460,
            "densidade_umida_g_cm3": 2.156,
            "umidade_media_percent": 9.8,
            "densidade_seca_g_cm3": 1.958,
        },
        {
            "ponto": 4,
            "peso_umido_molde_g": 9775,
            "peso_solo_umido_g": 4361,
            "densidade_umida_g_cm3": 2.109,
            "umidade_media_percent": 13.1,
            "densidade_seca_g_cm3": 1.865,
        },
    ]


def sample_data(language="pt", test_type="ambos"):
    return {
        "obra": "AV. BEIRA MAR",
        "interessado": "PREFEITURA MUNICIPAL",
        "empresa_executora": "RCA",
        "cidade": "PITIMBU",
        "procedencia": "AV. BEIRA MAR",
        "estaca": "ESTACA 02",
        "camada": "SUBLEITO",
        "lado": "D",
        "profundidade": "0,80m",
        "data_ensaio": "2026-07-08",
        "registro": "101",
        "responsavel_tecnico": "Eng. Responsável",
        "operador": "Operador A",
        "laboratorista": "Laboratorista A",
        "laboratorio": "Laboratório RCA",
        "proctor": "INTERMEDIÁRIO",
        "language": language,
        "test_type": test_type,
        "compactacao": {
            "umidade_otima": 9.8,
            "densidade_maxima": 1.958,
            "pontos": sample_compaction_points(),
        },
        "cbr": {"cbr": 34.0, "expansao": 0.5, "leituras": []},
        "granulometria": {
            "passante_10": 25.5,
            "passante_40": 19.7,
            "passante_200": 11.4,
            "classificacao_trb": "A-1-a",
            "classificacao_sucs": "GM",
            "limites": {"ll": "NL", "lp": "NP", "ip": "NP"},
        },
    }


@override_settings(MEDIA_ROOT=None)
def test_generate_workbooks_pt(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    files = generate_workbooks(1, sample_data("pt", "ambos"))
    assert len(files) == 2
    for item in files:
        validate_workbook(item["path"])
    compaction = next(item for item in files if "compactacao_cbr" in item["name"])
    wb = load_workbook(compaction["path"], data_only=False)
    dens = wb["DENS"]
    assert dens["B4"].value == "PREFEITURA MUNICIPAL"
    assert dens["F4"].value == "OBRA: AV. BEIRA MAR"
    assert dens["K4"].value == "INTERMEDIÁRIO"
    assert dens["M4"].value == "101"
    assert dens["B5"].value == "LOCAL: ESTACA 02"
    assert dens["B8"].value == "AV. BEIRA MAR"
    assert dens["F8"].value == "PITIMBU"
    assert dens["J8"].value == "Eng. Responsável"
    cbr = wb["CBR"]
    assert cbr["B4"].value == "PREFEITURA MUNICIPAL"
    assert cbr["F4"].value == "OBRA: AV. BEIRA MAR"
    assert cbr["I6"].value == "ESTACA 02"
    assert cbr["M6"].value == "0,80m"
    assert cbr["H17"].value == '2½"'

    granulometry = next(item for item in files if "granulometria" in item["name"])
    wb = load_workbook(granulometry["path"], data_only=False)
    gran = wb["GRANULOMEDIA "]
    assert gran["A27"].value == "AV. BEIRA MAR"
    assert gran["D27"].value == "AV. BEIRA MAR"
    assert gran["H27"].value == "RCA"
    assert gran["B29"].value == "SUBLEITO"
    assert gran["G29"].value == "ESTACA 02"
    assert gran["J29"].value == "D"
    assert gran["L29"].value == "0,80m"
    assert gran["A31"].value == "Laboratório RCA"
    assert gran["C31"].value == "Operador A"
    assert gran["G31"].value == "Laboratorista A"


@override_settings(MEDIA_ROOT=None)
def test_generate_workbooks_en(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    files = generate_workbooks(2, sample_data("en", "ambos"))
    assert len(files) == 2
    assert all(item["name"].endswith("_en.xlsx") for item in files)
    for item in files:
        validate_workbook(item["path"])


@override_settings(MEDIA_ROOT=None)
def test_cbr_record_can_differ_from_proctor_record(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    data = sample_data("pt", "compactacao_cbr")
    data["proctor_admin"] = {"registro": "101"}
    data["cbr_admin"] = {
        "registro": "102-CBR",
        "molde_numero": "03",
        "peso_molde": "5437",
        "volume_molde": "3211",
        "numero_camadas": "6",
        "golpes_camada": "27",
        "peso_soquete": "4537",
        "espessura_disco": "2,5",
        "altura_cilindro": "115",
        "constante_prensa": "0,084",
    }
    files = generate_workbooks(3, data)
    compaction = files[0]
    wb = load_workbook(compaction["path"], data_only=False)
    assert wb["DENS"]["M4"].value == "101"
    cbr = wb["CBR"]
    assert cbr["O4"].value == "102-CBR"
    assert cbr["H11"].value == "03"
    assert cbr["H12"].value == 5437
    assert cbr["H13"].value == 3211
    assert cbr["H14"].value == 6
    assert cbr["H15"].value == 27
    assert cbr["H16"].value == 4537
    assert cbr["H17"].value == 2.5
    assert cbr["H18"].value == 115
    assert cbr["P22"].value == 0.084


@override_settings(MEDIA_ROOT=None)
def test_generation_accepts_numeric_strings_from_ai(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    data = sample_data("pt", "ambos")
    data["compactacao"]["umidade_otima"] = "10,4%"
    data["compactacao"]["densidade_maxima"] = "1,907 g/cm³"
    data["cbr"]["cbr"] = "33,2%"
    data["cbr"]["expansao"] = "4 %"
    data["granulometria"]["passante_10"] = "73%"
    data["granulometria"]["passante_40"] = "51,0"
    data["granulometria"]["passante_200"] = "14 %"
    files = generate_workbooks(4, data)
    assert len(files) == 2
    for item in files:
        validate_workbook(item["path"])


@override_settings(MEDIA_ROOT=None)
def test_compaction_cbr_uses_extracted_moisture_inputs_and_keeps_formulas(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    data = sample_data("pt", "compactacao_cbr")
    data["cbr"].update({
        "higroscopica": {
            "capsula": "31",
            "peso_bruto_umido": 101.7,
            "peso_bruto_seco": 101.2,
            "tara_capsula": 17.29,
        },
        "moldagem": {
            "capsula": "40",
            "peso_bruto_umido": 100.4,
            "peso_bruto_seco": 92.0,
            "tara_capsula": 17.36,
        },
        "peso_solo_umido_passando_peneira_4": 6.88,
        "calculo_moldagem": {
            "peso_solo_umido": 6000,
            "peso_retido_peneira_4": 404,
            "peso_passando_peneira_4": 5596,
            "agua_a_juntar": 581,
        },
        "verificacao_moldagem": {
            "peso_bruto_cp_umido": 9620,
        },
    })
    files = generate_workbooks(5, data)
    compaction = files[0]
    validate_workbook(compaction["path"])
    wb = load_workbook(compaction["path"], data_only=False)
    dens = wb["DENS"]
    cbr = wb["CBR"]

    assert dens["G10"].value == "31"
    assert dens["G11"].value == 101.7
    assert dens["G12"].value == 101.2
    assert dens["G13"].value == 17.29
    assert dens["G14"].value == 0.5
    assert dens["G15"].value == "=G12-G13"
    assert dens["G16"].value == "=G14/G15*100"

    assert cbr["O11"].value == "31"
    assert cbr["O12"].value == 101.7
    assert cbr["O13"].value == 101.2
    assert cbr["O14"].value == 17.29
    assert cbr["O15"].value == "=O12-O13"
    assert cbr["O17"].value == "=O15/O16*100"

    assert cbr["P11"].value == "40"
    assert cbr["P12"].value == 100.4
    assert cbr["P13"].value == 92
    assert cbr["P14"].value == 17.36
    assert cbr["P15"].value == "=P12-P13"
    assert cbr["P17"].value == "=P15/P16*100"

    assert cbr["F21"].value == 9.8
    assert cbr["F22"].value == "=O17"
    assert cbr["F23"].value == "=F21-F22"
    assert cbr["M20"].value == 6.88
    assert cbr["M21"].value == "=M20/(1+(O18/100))"
    assert cbr["O21"].value == '=IF(OR(M21="",F23=""),"",M21*F23/100)'
    assert cbr["P36"].value == 6000
    assert cbr["P38"].value == 404
    assert cbr["P40"].value == 5596
    assert cbr["P42"].value == 581
    assert cbr["P45"].value == 9.62
    assert round(cbr["P47"].value, 3) == 4.184
    assert cbr["P49"].value == "=P47/H13*1000"
    assert cbr["P51"].value == "=P49/(1+(P18/100))"


@override_settings(MEDIA_ROOT=None)
def test_invalid_extracted_moisture_block_is_not_written(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    data = sample_data("pt", "compactacao_cbr")
    data["cbr"]["higroscopica"] = {
        "capsula": "31",
        "peso_bruto_umido": 101.7,
        "peso_bruto_seco": 102.2,
        "tara_capsula": 17.29,
    }
    files = generate_workbooks(6, data)
    compaction = files[0]
    validate_workbook(compaction["path"])
    wb = load_workbook(compaction["path"], data_only=False)
    dens = wb["DENS"]
    cbr = wb["CBR"]

    assert dens["G10"].value == ""
    assert dens["G11"].value is None
    assert dens["G14"].value is None
    assert cbr["O11"].value == ""
    assert cbr["O12"].value is None
    assert cbr["O15"].value == "=O12-O13"


@override_settings(MEDIA_ROOT=None)
def test_compaction_hygroscopic_moisture_fills_cbr_when_cbr_block_is_invalid(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    data = sample_data("pt", "compactacao_cbr")
    data["compactacao"]["higroscopica"] = {
        "capsula": "31",
        "peso_bruto_umido": 101.7,
        "peso_bruto_seco": 101.2,
        "tara_capsula": 17.29,
    }
    data["cbr"]["higroscopica"] = {
        "capsula": "40",
        "peso_bruto_umido": 101.5,
        "peso_bruto_seco": 104.2,
        "tara_capsula": 17.99,
    }

    files = generate_workbooks(8, data)
    compaction = files[0]
    validate_workbook(compaction["path"])
    wb = load_workbook(compaction["path"], data_only=False)
    dens = wb["DENS"]
    cbr = wb["CBR"]

    assert dens["G10"].value == "31"
    assert dens["G11"].value == 101.7
    assert dens["G12"].value == 101.2
    assert dens["G13"].value == 17.29
    assert cbr["O11"].value == "31"
    assert cbr["O12"].value == 101.7
    assert cbr["O13"].value == 101.2
    assert cbr["O14"].value == 17.29


@override_settings(MEDIA_ROOT=None)
def test_cbr_molding_defaults_are_not_invented(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    data = sample_data("pt", "compactacao_cbr")
    files = generate_workbooks(7, data)
    compaction = files[0]
    validate_workbook(compaction["path"])
    wb = load_workbook(compaction["path"], data_only=False)
    cbr = wb["CBR"]

    assert cbr["M20"].value is None
    assert cbr["M21"].value == "=M20/(1+(O18/100))"
    assert cbr["P36"].value is None
    assert cbr["P38"].value is None
    assert cbr["P40"].value is None
    assert cbr["P42"].value is None
    assert cbr["P45"].value is None
    assert cbr["P47"].value is None
    assert cbr["P49"].value == "=P47/H13*1000"

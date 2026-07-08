from __future__ import annotations

import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def validate_zip(path: Path) -> None:
    with zipfile.ZipFile(path) as workbook:
        bad_file = workbook.testzip()
        if bad_file:
            raise RuntimeError(f"{path}: arquivo interno corrompido: {bad_file}")


def validate_xml(path: Path) -> None:
    with zipfile.ZipFile(path) as workbook:
        for name in workbook.namelist():
            if name.endswith(".xml") or name.endswith(".rels"):
                ET.fromstring(workbook.read(name))


def validate_openpyxl(path: Path) -> list[str]:
    workbook = load_workbook(path, data_only=False, read_only=False, keep_links=True)
    return workbook.sheetnames


def main() -> None:
    args = [Path(arg).resolve() for arg in sys.argv[1:]]
    files = args or sorted((ROOT / "templates").glob("*.xlsx")) + sorted((ROOT / "output").glob("*.xlsx"))
    if not files:
        print("Nenhuma planilha encontrada para validar.")
        return

    for path in files:
        validate_zip(path)
        validate_xml(path)
        sheetnames = validate_openpyxl(path)
        try:
            label = path.relative_to(ROOT)
        except ValueError:
            label = path
        print(f"OK {label} :: {', '.join(sheetnames)}")


if __name__ == "__main__":
    main()

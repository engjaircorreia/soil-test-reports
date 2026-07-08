import pytest
from django.test import override_settings
from openpyxl import load_workbook

from ensaios.services.workbook_fill import generate_workbooks


def minimal_reviewed_data():
    return {
        "language": "pt",
        "test_type": "compactacao_cbr",
        "proctor_admin": {
            "interessado": "Cliente",
            "obra": "Obra",
            "proctor": "INTERMEDIARIO",
            "estaca": "E02",
            "procedencia": "Rua",
            "data_ensaio": "2026-07-08",
        },
        "cbr_admin": {
            "registro": "1",
            "molde_numero": "02",
            "peso_molde": "4775",
            "volume_molde": "2088",
            "numero_camadas": "5",
            "golpes_camada": "26",
            "peso_soquete": "4536",
            "espessura_disco": '2 1/2"',
            "altura_cilindro": "114",
            "constante_prensa": "0,0839",
        },
        "compactacao": {
            "umidade_otima": 10.0,
            "densidade_maxima": 1.96,
            "pontos": [],
        },
        "cbr": {
            "cbr": 39.0,
            "expansao": 0.7,
            "higroscopica": {
                "capsula": "40",
                "peso_bruto_umido": 101.5,
                "peso_bruto_seco": 101.2,
                "tara_capsula": 17.99,
            },
            "moldagem": {
                "capsula": "32",
                "peso_bruto_umido": 100.3,
                "peso_bruto_seco": 92.3,
                "tara_capsula": 17.35,
            },
            "leituras": [],
        },
    }


@override_settings(MEDIA_ROOT=None)
def test_workbook_generation_blocks_when_required_raw_proctor_points_are_missing(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path

    with pytest.raises(ValueError, match="pontos"):
        generate_workbooks(101, minimal_reviewed_data())


@override_settings(MEDIA_ROOT=None)
def test_workbook_generation_does_not_synthesize_proctor_table_from_final_result(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    data = minimal_reviewed_data()
    data["compactacao"]["pontos"] = [
        {
            "ponto": 1,
            "peso_umido_molde_g": 9374,
            "peso_solo_umido_g": 3960,
            "umidade_media_percent": 3.5,
            "densidade_seca_g_cm3": 1.85,
        }
    ]

    files = generate_workbooks(102, data)
    workbook_path = files[0]["path"]
    wb = load_workbook(workbook_path, data_only=False)
    dens = wb["DENS"]

    assert dens["C23"].value == 9374
    assert dens["D23"].value == 3960
    assert dens["L23"].value == 3.5
    assert dens["N23"].value == 1.85
    assert dens["C24"].value in (None, "")


@override_settings(MEDIA_ROOT=None)
def test_workbook_generation_does_not_synthesize_cbr_penetration_from_final_result(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    data = minimal_reviewed_data()
    data["compactacao"]["pontos"] = [
        {
            "ponto": 1,
            "peso_umido_molde_g": 9374,
            "peso_solo_umido_g": 3960,
            "umidade_media_percent": 3.5,
            "densidade_seca_g_cm3": 1.85,
        }
    ]
    data["cbr"]["leituras"] = [
        {"linha": 29, "leitura_extensometro": 282, "pressao_corrigida_kg_cm2": 273.5, "cbr_percent": 39.0},
        {"linha": 30, "leitura_extensometro": 338, "pressao_corrigida_kg_cm2": 327.8, "cbr_percent": 31.0},
    ]

    files = generate_workbooks(103, data)
    workbook_path = files[0]["path"]
    wb = load_workbook(workbook_path, data_only=False)
    cbr = wb["CBR"]

    assert cbr["F29"].value == 282
    assert cbr["G29"].value == 273.5
    assert cbr["J29"].value == 39.0
    assert cbr["F30"].value == 338
    assert cbr["G30"].value == 327.8
    assert cbr["J30"].value == 31.0
